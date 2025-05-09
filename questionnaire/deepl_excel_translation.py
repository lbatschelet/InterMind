#!/usr/bin/env python3
"""
translate_excel_deepl.py
========================
CLI translator for questionnaire workbooks using DeepL’s *document* API.

Workflow
--------
1. **Validation** (unless --no-validate)  
   - ensures every choice question has the same number of labels/values  
   - slider questions have 2–3 labels
2. For each target language:  
   a. Copy the source sheet into a tiny temp workbook  
   b. Call DeepL `translate_document_from_filepath()`  
   c. Insert the translated sheet into the full workbook (name = target code)

By default the script writes `<file>_translated.xlsx`.  
Pass `--in-place` to overwrite the original file.

Minimal example
---------------
# New file with English & French sheets
python translate_excel_deepl.py survey.xlsx --src-lang DE --targets EN-US FR

# Overwrite original & skip validation
python translate_excel_deepl.py survey.xlsx --src-lang DE --targets IT --in-place --no-validate

Flags
-----
--src-lang      DeepL source language (required)
--src-sheet     Sheet name to translate            [default: de]
--targets       One or more DeepL target codes     (required)
--formality     default | more | less | prefer_more | prefer_less
--glossary      DeepL glossary id
--in-place      Overwrite workbook instead of writing *_translated.xlsx
--no-validate   Skip consistency checks
--api-key       Override DEEPL_API_KEY environment variable

DeepL document-API reference → https://developers.deepl.com/docs/api-reference/document/openapi-spec-for-document-translation
"""

import argparse, os, sys, time, shutil, tempfile, re
from pathlib import Path

import deepl
import openpyxl
import pandas as pd

MAX_OPTS = 20
SLIDER_MIN, SLIDER_MAX = 2, 3
LANG_RE = re.compile(r"^[a-z]{2}(?:-[a-z]{2})?$", re.I)


# ─────────────── validation helpers ────────────────────────────────────────
def _norm(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    return df


def _opt_vals(row) -> list[str]:
    return [
        str(row[f"opt_val_{i}"]).strip()
        for i in range(1, MAX_OPTS + 1)
        if f"opt_val_{i}" in row and pd.notna(row[f"opt_val_{i}"]) and str(row[f"opt_val_{i}"]).strip()
    ]


def _labels(row, prefix: str, limit: int) -> list[str]:
    return [
        str(row[f"{prefix}{i}"]).strip()
        for i in range(1, limit + 1)
        if f"{prefix}{i}" in row and pd.notna(row[f"{prefix}{i}"]) and str(row[f"{prefix}{i}"]).strip()
    ]


def validate(wb: Path, src_sheet: str) -> None:
    xl = pd.ExcelFile(wb)
    if "base" not in xl.sheet_names:
        raise ValueError("Workbook is missing a 'base' sheet.")
    if src_sheet not in xl.sheet_names:
        raise ValueError(f"Sheet '{src_sheet}' not found.")

    base = _norm(xl.parse("base"))
    src = _norm(xl.parse(src_sheet))
    if "question_id" in base.columns and "id" not in base.columns:
        base.rename(columns={"question_id": "id"}, inplace=True)

    for _, tr in src.iterrows():
        qid = tr["question_id"]
        brow = base.loc[base["id"] == qid]
        if brow.empty:
            raise ValueError(f"unknown question_id '{qid}' in sheet '{src_sheet}'")
        brow = brow.iloc[0]

        qtype = brow["type"]
        if qtype in {"single_choice", "multiple_choice"}:
            if len(_labels(tr, "opt_label_", MAX_OPTS)) != len(_opt_vals(brow)):
                raise ValueError(f"label/value count mismatch for '{qid}' in '{src_sheet}'")
        elif qtype == "slider":
            n = len([v for v in _labels(tr, "slider_val_", SLIDER_MAX) if v])
            if not (SLIDER_MIN <= n <= SLIDER_MAX):
                raise ValueError(f"slider '{qid}' needs {SLIDER_MIN}-{SLIDER_MAX} labels (found {n})")


# ─────────────── DeepL + sheet copy helpers ───────────────────────────────
def wait(handle: deepl.DocumentHandle, tx: deepl.Translator):
    while True:
        st = tx.document_status(handle)
        if st.status == "done":
            return
        if st.status == "error":
            raise RuntimeError(f"DeepL error: {st.error_message}")
        time.sleep(min(5, st.seconds_remaining or 5))


def copy_sheet(src_sheet, dest_wb, dest_title: str):
    target = dest_wb.create_sheet(dest_title)
    for ri, row in enumerate(src_sheet.iter_rows(values_only=True), 1):
        for ci, val in enumerate(row, 1):
            target.cell(ri, ci, val)


# ─────────────── CLI / main ───────────────────────────────────────────────
def cli() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Translate one sheet via DeepL document API.")
    p.add_argument("workbook", help=".xlsx questionnaire")
    p.add_argument("--src-lang", required=True, help="source language code (DE, EN …)")
    p.add_argument("--src-sheet", default="de", help="sheet to translate [de]")
    p.add_argument("--targets", required=True, nargs="+", help="target codes EN-US FR …")
    p.add_argument("--formality", default="default",
                   choices=["default", "more", "less", "prefer_more", "prefer_less"])
    p.add_argument("--glossary", default=None, help="DeepL glossary id")
    p.add_argument("--in-place", action="store_true", help="overwrite workbook")
    p.add_argument("--no-validate", action="store_true", help="skip validation")
    p.add_argument("--api-key", default=os.getenv("DEEPL_API_KEY"),
                   help="override env var")
    return p.parse_args()


def main():
    args = cli()
    if not args.api_key:
        sys.exit("DEEPL_API_KEY missing (env var or --api-key).")

    src_file = Path(args.workbook).resolve()
    if not src_file.exists():
        sys.exit(f"File not found: {src_file}")

    if not args.no_validate:
        try:
            validate(src_file, args.src_sheet)
            print("✓ validation passed")
        except ValueError as e:
            sys.exit(f"Validation failed: {e}")

    out_file = src_file if args.in_place else src_file.with_stem(src_file.stem + "_translated")
    if out_file != src_file:
        shutil.copy(src_file, out_file)

    tx = deepl.Translator(args.api_key)

    for tgt in args.targets:
        print(f"→ translating {args.src_sheet} → {tgt}")
        # 1. build temp workbook with only source sheet
        temp_src = Path(tempfile.mkstemp(suffix=".xlsx")[1])
        wb_full = openpyxl.load_workbook(out_file)
        wb_tmp = openpyxl.Workbook()
        copy_sheet(wb_full[args.src_sheet], wb_tmp, args.src_sheet)
        wb_tmp.remove(wb_tmp["Sheet"])
        wb_tmp.save(temp_src)

        # 2. DeepL translates the whole temp file
        temp_tr = temp_src.with_stem(temp_src.stem + "_t")
        tx.translate_document_from_filepath(
            str(temp_src), str(temp_tr),
            source_lang=args.src_lang,
            target_lang=tgt,
            formality=None if args.formality == "default" else args.formality,
            glossary=args.glossary,
        )

        # 3. merge translated sheet back
        wb_dest = openpyxl.load_workbook(out_file)
        wb_tran = openpyxl.load_workbook(temp_tr, data_only=True)

        # try exact match, then case-insensitive, then simply first sheet
        if args.src_sheet in wb_tran.sheetnames:
            src_ws = wb_tran[args.src_sheet]
        else:
            alt = next((s for s in wb_tran.sheetnames
                        if s.lower() == args.src_sheet.lower()), wb_tran.sheetnames[0])
            src_ws = wb_tran[alt]

        copy_sheet(src_ws, wb_dest, tgt.lower())
        wb_dest.save(out_file)

        temp_src.unlink(); temp_tr.unlink()
        print(f"✓ added sheet '{tgt.lower()}'")

    print("Done →", out_file)


if __name__ == "__main__":
    main()
